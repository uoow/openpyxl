import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

wb = openpyxl.Workbook()

# Font
ws = wb.active
ws.title = 'Font'

# default 11pt, Calibri
italic24Font = Font(size=24,italic=True)
ws['B3'].font = italic24Font
ws['B3'] = '24pt Italic'

boldRedFont = Font(name='Times New Roman', bold=True,color=colors.RED)
ws['A1'].font = boldRedFont
ws['A1'] = 'Bold Red Times New Roman'


# Formulas
ws = wb.create_sheet('Formula')
ws['A1'] = 200
ws['A2'] = 300
ws['A3'] = '=SUM(A1:A2)'

# Setting row height and column width
ws = wb.create_sheet('dimensions')
ws['A1'] = 'Tall row'
ws.row_dimensions[1].height = 70
ws['B2'] = 'Wide column'
ws.column_dimensions['B'].width = 20


# Merging cells
ws = wb.create_sheet('merged')
ws.merge_cells('A1:D3')
ws['A1'] = 'Twelve cells merged together'
ws.merge_cells('C5:D5')
ws['C5'] = 'Twelve cells merged together'


# Unmerging cells
ws = wb.copy_worksheet(wb['merged'])
ws.title = 'unmerged'
ws.unmerge_cells('A1:D3')
ws.unmerge_cells('C5:D5')

wb.save('style.xlsx')


# chart

from openpyxl.chart import(
    Reference,
    Series,
    PieChart,
    BarChart,
    BubbleChart
)

wb = openpyxl.Workbook()

# Pie chart
ws = wb.active
ws.title = 'pieChart'

data = [
    ['Pie','Sold'],
    ['Apple',50],
    ['Cherry',30],
    ['Pumpkin',10],
    ['Chocolate',40],
]
for row in data:
    ws.append(row)

pie = PieChart()
labels = Reference(ws,min_col=1,min_row=2,max_row=5)
data = Reference(ws,min_col=2,min_row=2,max_row=5)
pie.add_data(data)
pie.set_categories(labels)
pie.title = 'Pie sold by category'
ws.add_chart(pie,'A15')


# Bar chart
ws = wb.create_sheet('barChart')

rows = [
    ('Number','Batch 1','Batch 2'),
    (2,10,30),
    (3,40,60),
    (4,50,70),
    (5,20,10),
    (6,10,40),
    (7,50,30),
]

for row in rows:
    ws.append(row)

chart1 = BarChart()
chart1.type = 'col'
chart1.style = 10
chart1.title = 'Bar Chart'
chart1.y_axis.title = 'Sample length(mm)'
chart1.x_axis.title = 'Test number'

cats = Reference(ws,min_col=1,min_row=2,max_row=7)
data = Reference(ws,min_col=2,max_col=3,min_row=2,max_row=7)
chart1.add_data(data)
chart1.set_categories(cats)
ws.add_chart(chart1,'A10')


# Bubble chart
ws = wb.create_sheet('bubbleChart')

rows = [
    ("Number of Products","Sales in USD","Market share"),
    (14,12200,15),
    (20,60000,33),
    (18,24400,10),
    (22,32000,42),
    (),
    (12,8200,18),
    (15,50000,30),
    (19,22400,15),
    (25,25000,50),
]

for row in rows:
    ws.append(row)

chart = BubbleChart()
chart.style = 18

# add the first series of data
xvalues = Reference(ws,min_col=1,min_row=2,max_row=5)
yvalues = Reference(ws,min_col=2,min_row=2,max_row=5)
size = Reference(ws,min_col=3,min_row=2,max_row=5)
series = Series(values=yvalues,xvalues=xvalues,zvalues=size,title='2013')
chart.series.append(series)

# add the second series of data
xvalues2 = Reference(ws,min_col=1,min_row=7,max_row=10)
yvalues2 = Reference(ws,min_col=2,min_row=7,max_row=10)
size2 = Reference(ws,min_col=3,min_row=7,max_row=10)
series2 = Series(values=yvalues2,xvalues=xvalues2,zvalues=size2,title='2014')
chart.series.append(series2)

# place the chart starting in cell E1
ws.add_chart(chart,'E1')

wb.save('chart.xlsx')